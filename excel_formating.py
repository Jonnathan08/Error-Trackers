from openpyxl import Workbook, load_workbook
from openpyxl.styles import colors , fills, Font
from openpyxl.utils import get_column_letter
from datetime import datetime

def format_excel():

    today = datetime.strftime(datetime.today(), '%m-%d-%Y')

    wb = load_workbook('error_report_%s.xlsx' % today)

    ws = wb.active

    orange = colors.Color(rgb='FABF8F')
    filling = fills.PatternFill(patternType='solid', fgColor=orange)
    light_orange = colors.Color(rgb='FCD5B4')
    filling2 = fills.PatternFill(patternType='solid', fgColor=light_orange)
    black = colors.Color(rgb='000000')
    filling_title = fills.PatternFill(patternType='solid', fgColor=black)
    ft = Font(color="FFFFFF") # blanco

    #add colors in columns
    for col in range (1, ws.max_column+1):
        for cel in range (1, len(ws['A']) + 1):        
            if col % 2 != 0:
                ws[get_column_letter(col) + str(cel)].fill = filling
            else:
                ws[get_column_letter(col) + str(cel)].fill = filling2

            if cel == 1:
                ws[get_column_letter(col) + str(cel)].fill = filling_title
                ws[get_column_letter(col) + str(cel)].font = ft

    #adjust columns width
    for i in range(1, ws.max_column+1):
        if i < 3:
            ws.column_dimensions[get_column_letter(i)].width = 20
        else:
            ws.column_dimensions[get_column_letter(i)].width = 30

    #add filter in every column
    ws.auto_filter.ref = ws.dimensions


    wb.save('error_report_%s.xlsx' % today)

#format_excel()