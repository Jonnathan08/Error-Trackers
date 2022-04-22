from openpyxl import load_workbook
from openpyxl.styles import colors , fills
from openpyxl.utils import get_column_letter
from datetime import datetime

def format_excel():

    today = datetime.strftime(datetime.today(), '%m-%d-%Y')

    wb = load_workbook('compass_error_report_%s.xlsx' % today)

    ws = wb.active

    blue_dark = colors.Color(rgb='BDD7EE')
    filling_title_1 = fills.PatternFill(patternType='solid', fgColor=blue_dark)
    blue_light = colors.Color(rgb='DDEBF7')
    filling_1 = fills.PatternFill(patternType='solid', fgColor=blue_light)
    orange_dark = colors.Color(rgb='F8CBAD')
    filling_title_2 = fills.PatternFill(patternType='solid', fgColor=orange_dark)
    orange_light = colors.Color(rgb='FCE4D6')
    filling_2 = fills.PatternFill(patternType='solid', fgColor=orange_light)

    #add colors in columns
    for col in range (1, ws.max_column+1):
        for cel in range (1, len(ws['A']) + 1):       
            if get_column_letter(col) in ['A','B','C','D','E','F','G','H']:
                if cel == 1:
                    ws[get_column_letter(col) + str(cel)].fill = filling_title_1
                else:
                    ws[get_column_letter(col) + str(cel)].fill = filling_1
            else:
                if cel == 1:
                    ws[get_column_letter(col) + str(cel)].fill = filling_title_2
                else:
                    ws[get_column_letter(col) + str(cel)].fill = filling_2

    #adjust columns width
    for i in range(2, ws.max_column+1):
        if i < 3:
            ws.column_dimensions[get_column_letter(i)].width = 20
        else:
            ws.column_dimensions[get_column_letter(i)].width = 30

    #add filter in every column
    ws.auto_filter.ref = ws.dimensions


    wb.save('compass_error_report_%s.xlsx' % today)

#format_excel()